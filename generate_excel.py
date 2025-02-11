from faker import Faker  # Libreria che genera dati casuali
import openpyxl  # Libreria per scrittura e lettura del file excel
from openpyxl.styles import Font  # Importa classe Font per cambiare stile Font nelle celle di Excel
from openpyxl.utils import get_column_letter  # Funzione che consente di convertire gli indici delle colonne in lettere
import os

# Funzione per generare dati casuali degli studenti
def generate_student_data(num_students):
    fake = Faker('it_IT')  # Crea oggetto che consente generare dati in italiano
    student_data = []  # Inizializzazione lista vuota
     
    for _ in range(num_students):
        nome = fake.first_name()
        cognome = fake.last_name()
        data_di_nascita = fake.date_of_birth(minimum_age=18, maximum_age=35).strftime('%d/%m/%Y')  # crea una data di nascita per utenti tra 18 e 35 anni e la setta nel formato gg/mm/aaaa
        luogo_di_nascita = fake.city()
        codice_fiscale = fake.ssn()
        email = fake.email()
        contatto_telefonico = fake.phone_number()
        indirizzo_residenza = fake.address()
        
        student_data.append([nome, cognome, data_di_nascita, luogo_di_nascita, codice_fiscale, email, contatto_telefonico, indirizzo_residenza])  # Aggiunge dati generati alla lista student_data
    
    return student_data  # Restituisce lista student_data

# Funzione per scrivere dati nel file Excel
def write_to_excel(file_name, headers, data):
    wb = openpyxl.Workbook()  # Crea workbook Excel
    ws = wb.active  # Attiva foglio di lavoro
    ws.title = "Utenti" 

    ws.append(headers)

    #  Ciclo che scrive l'intestazione nella prima riga e imposta il font in grassetto 
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  

    for row in data:
        ws.append(row)

    # Ottiene/Crea percorso directory
    output_directory = os.path.dirname(file_name)
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Regola la larghezza delle colonne in base ai dati
    for col in ws.iter_cols(): #  Itera le colonne
        max_length = 0 
        column = get_column_letter(col[0].column)  # Ottiene lettera della colonna basandosi su indice di colonna
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2 # Da più spazio alla cella
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_name)
    print(f"Il file Excel '{file_name}' è stato generato correttamente!")

if __name__ == "__main__":
     # Ottiene il percorso della directory corrente
    current_directory = os.path.dirname(__file__)

    # Costruisce il percorso del file dinamicamente
    file_name = os.path.join(current_directory, 'utenti.xlsx')
    
    num_students = 10
    headers = ["NOME", "COGNOME", "DATA DI NASCITA", "LUOGO DI NASCITA", "CODICE FISCALE", "EMAIL", "CONTATTO TELEFONICO", "RESIDENZA\\INDIRIZZO"]
    student_data = generate_student_data(num_students)

    write_to_excel(file_name, headers, student_data)
