import os
import sqlite3
import pandas as pd
from openpyxl.utils import get_column_letter
from generate_db import excel_to_sql
import threading
import logging

# Configura il logger per il debug
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Funzione per esportare i dati dal database SQLite a un file Excel
def sql_to_excel(db_file, excel_file):
    conn = None
    try:
        logging.debug("Tentativo di connessione al database SQLite...")
        conn = sqlite3.connect(db_file)  # Connessione al database SQLite
        logging.debug("Connessione riuscita.")
        
        # Esecuzione della query SQL per ottenere tutti i dati dalla tabella 'utenti'
        df = pd.read_sql_query('SELECT * FROM utenti', conn)
        logging.debug("Dati ottenuti dal database.")

        # Scrittura dei dati su un file Excel utilizzando openpyxl
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Utenti')
            ws = writer.sheets['Utenti']

            # Regola la larghezza delle colonne in base ai dati
            for col in ws.iter_cols():
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.error(f"Errore nel calcolo della larghezza della colonna: {e}")
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

        logging.debug("File Excel aggiornato con successo.")
    except sqlite3.Error as e:
        logging.error(f"Errore durante l'interazione con il database SQLite: {e}")
    except Exception as e:
        logging.error(f"Errore durante la scrittura del file Excel: {e}")
    finally:
        if conn:
            conn.close()  # Chiusura della connessione al database SQLite

# Funzione per sincronizzare i file Excel e il database SQLite
def synchronize_files(excel_file, db_file):
    try:
        logging.debug("Tentativo di sincronizzazione dei file...")
        
        # Se il database non esiste, viene creato dai dati nel file Excel
        if not os.path.exists(db_file):
            excel_to_sql(excel_file, db_file)
            logging.debug(f"Database '{db_file}' creato e aggiornato dai dati Excel.")

        # Ottieni il tempo di modifica dei file
        last_excel_mod_time = os.path.getmtime(excel_file)
        last_db_mod_time = os.path.getmtime(db_file)

        # Confronta i tempi di modifica e aggiorna se necessario
        if last_excel_mod_time > last_db_mod_time:
            logging.debug("Il file Excel è stato aggiornato. Sincronizzazione del database...")
            excel_to_sql(excel_file, db_file)
            logging.debug(f"Database '{db_file}' aggiornato con i dati dal file Excel.")
        else:
            logging.debug("Il database è già aggiornato rispetto al file Excel.")

    except FileNotFoundError as e:
        logging.error(f"Errore: File non trovato - {e}")
    except sqlite3.Error as e:
        logging.error(f"Errore durante l'accesso o la sincronizzazione del database: {e}")
    except Exception as e:
        logging.error(f"Errore generale durante la sincronizzazione: {e}")

# Funzione per eseguire la sincronizzazione in un thread separato per evitare che la GUI si blocchi
def synchronize_files_thread(excel_file, db_file):
    threading.Thread(target=synchronize_files, args=(excel_file, db_file), daemon=True).start()


